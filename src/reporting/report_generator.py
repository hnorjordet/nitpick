"""
Excel QA report generator — produces an Xbench-style .xlsx report.
"""
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Colour palette ────────────────────────────────────────────────────────────
_HEADER_FILL   = PatternFill("solid", fgColor="2D5FA6")   # dark blue
_SECTION_FILL  = PatternFill("solid", fgColor="4A7FCC")   # mid blue
_ALT_FILL      = PatternFill("solid", fgColor="EEF2F9")   # light blue tint
_ERROR_FILL    = PatternFill("solid", fgColor="FFECEC")   # light red
_WARN_FILL     = PatternFill("solid", fgColor="FFF8E1")   # light yellow
_WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

_SECTION_TYPES = {
    # check_source → section label
    "termlist": "Terminology",
    "checklist": "Checklist",
    "number": "Number / Placeholder",
    "qa": "QA Checks",
    "spell": "Spellcheck",
}

_QA_SEVERITY = {
    "untranslated": "error",
    "source_equals_target": "warning",
    "inconsistent_source": "warning",
    "inconsistent_target": "warning",
    "tag_mismatch": "error",
    "url_email_mismatch": "warning",
    "alphanumeric_mismatch": "warning",
    "double_blanks": "info",
    "repeated_words": "info",
    "uppercase_mismatch": "warning",
    "camelcase_mismatch": "info",
    "number_mismatch": "error",
    "placeholder_mismatch": "error",
    "dnt_translated": "error",
    "missing_required": "error",
    "forbidden_found": "error",
    "rule_violation": "warning",
    "unpaired_symbol": "warning",
}


def _cell_style(ws, row: int, col: int,
                value=None, bold=False, wrap=True,
                fill=None, font_color="000000",
                halign="left", valign="top",
                font_size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=font_size)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                                wrap_text=wrap)
    cell.border = _THIN
    if fill:
        cell.fill = fill
    return cell


def _section_header(ws, row: int, label: str, count: int, ncols: int,
                    subtitle: str = ""):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    suffix = f"  —  from {subtitle}" if subtitle else ""
    cell = ws.cell(row=row, column=1,
                   value=f"  {label}{suffix}  —  {count} issue{'s' if count != 1 else ''}")
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = _SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1)
    cell.border = _THIN


def generate_xlsx_report(
    file_path: str,
    spell_errors: list,   # list of {"word": str, "count": int, "segment_ids": list[str]}
    violations: list,     # list of Violation dicts
    output_path: str,
    app_version: str = "",
) -> str:
    """
    Build an Xbench-style Excel QA report and write it to output_path.
    Returns the output_path on success.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "QA Report"

    file_name = Path(file_path).name
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    ncols = 8   # Seg #, File, Match %, Type, Source, Target, Description, Corrected

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [9, 28, 9, 20, 40, 40, 35, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    title = f"Nitpick QA Report — v{app_version}" if app_version else "Nitpick QA Report"
    title_cell = ws.cell(row=row, column=1, value=title)
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = _HEADER_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                      indent=1)
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    meta_cell = ws.cell(row=row, column=1,
                        value=f"File: {file_name}    Generated: {generated}")
    meta_cell.font = Font(size=9, color="FFFFFF", italic=True)
    meta_cell.fill = _HEADER_FILL
    meta_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
    ws.row_dimensions[row].height = 16
    row += 2  # blank row

    # ── Summary ───────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value="Summary")
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = _SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    row += 1

    spell_count = sum(e["count"] for e in spell_errors)
    viol_count = len(violations)
    summary_rows = [
        ("Spelling errors", spell_count),
        ("Terminology / QA violations", viol_count),
        ("Total issues", spell_count + viol_count),
    ]
    for label, cnt in summary_rows:
        _cell_style(ws, row, 1, label, bold=True, fill=_WHITE_FILL)
        c = _cell_style(ws, row, 2, cnt, halign="right", fill=_WHITE_FILL)
        c.font = Font(bold=True, size=10,
                      color="CC0000" if cnt > 0 else "338844")
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=ncols)
        row += 1

    row += 1  # blank

    # ── Column headers ────────────────────────────────────────────────────────
    headers = ["Seg #", "File", "Match %", "Type", "Source", "Target", "Description", "Corrected"]
    for col_i, hdr in enumerate(headers, 1):
        _cell_style(ws, row, col_i, hdr, bold=True,
                    fill=_HEADER_FILL, font_color="FFFFFF",
                    halign="center", valign="center")
        ws.row_dimensions[row].height = 18
    row += 1
    col_header_row = row  # remember for freeze

    # ── Spelling errors section ───────────────────────────────────────────────
    if spell_errors:
        _section_header(ws, row, "Spellcheck", len(spell_errors), ncols)
        row += 1

        for alt_i, entry in enumerate(spell_errors):
            fill = _ALT_FILL if alt_i % 2 == 0 else _WHITE_FILL
            word = entry.get("word", "")
            count = entry.get("count", 0)
            seg_ids = ", ".join(str(s) for s in entry.get("segment_ids", []))
            _cell_style(ws, row, 1, seg_ids,           fill=fill)
            _cell_style(ws, row, 2, file_name,         fill=fill)
            _cell_style(ws, row, 3, "",                fill=fill)  # no match% for spell
            _cell_style(ws, row, 4, "Spelling",        fill=fill)
            _cell_style(ws, row, 5, word,              fill=fill)
            _cell_style(ws, row, 6, "",                fill=fill)
            _cell_style(ws, row, 7, f"Flagged {count}×", fill=fill)
            _cell_style(ws, row, 8, "",                fill=fill)  # Corrected (editable)
            row += 1

    # ── Violations by source group ────────────────────────────────────────────
    def _viol_group(source_keys):
        return [v for v in violations
                if v.get("check_source", "qa") in source_keys]

    def _group_subtitle(source_keys):
        """Return a filename subtitle for termlist/checklist groups."""
        for v in violations:
            if v.get("check_source") in source_keys:
                # source_file is populated by term_checker when available
                sf = v.get("source_file", "")
                if sf:
                    return Path(sf).name
        return ""

    groups = [
        ("termlist",  ["termlist"]),
        ("checklist", ["checklist"]),
        ("number",    ["number"]),
        ("qa",        ["qa"]),
    ]

    for group_key, source_keys in groups:
        grp = _viol_group(source_keys)
        if not grp:
            continue
        label = _SECTION_TYPES.get(group_key, group_key.title())
        subtitle = _group_subtitle(source_keys) if group_key in ("termlist", "checklist") else ""
        _section_header(ws, row, label, len(grp), ncols, subtitle=subtitle)
        row += 1

        for alt_i, v in enumerate(grp):
            severity = _QA_SEVERITY.get(v.get("violation_type", ""), "warning")
            fill = _ERROR_FILL if severity == "error" else (
                   _WARN_FILL if severity == "warning" else _ALT_FILL
                   if alt_i % 2 == 0 else _WHITE_FILL)
            match_pct = v.get("match_percent")
            _cell_style(ws, row, 1, v.get("segment_id", ""),    fill=fill)
            _cell_style(ws, row, 2, v.get("file_name", ""),     fill=fill)
            _cell_style(ws, row, 3, match_pct,                  fill=fill, halign="center")
            _cell_style(ws, row, 4, v.get("violation_type", "").replace("_", " ").title(), fill=fill)
            _cell_style(ws, row, 5, v.get("source_text", ""),   fill=fill)
            _cell_style(ws, row, 6, v.get("target_text", ""),   fill=fill)
            _cell_style(ws, row, 7, v.get("description", ""),   fill=fill)
            corrected_cell = _cell_style(ws, row, 8, "",        fill=_WHITE_FILL)
            corrected_cell.protection = corrected_cell.protection  # stays unlocked
            row += 1

    # ── Freeze panes below header ─────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=col_header_row, column=1)

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A{col_header_row - 1}:{get_column_letter(ncols)}{row - 1}"
    )

    wb.save(output_path)
    return output_path
